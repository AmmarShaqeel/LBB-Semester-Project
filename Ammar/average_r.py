import openpyxl
import sys

#file_name = input("What's the file name? ")
file_name = sys.argv[1]

wb = openpyxl.load_workbook(file_name)
type(wb)

sheets = wb.get_sheet_names()

data_sheet = wb.get_sheet_by_name(sheets[0])
new_sheet = wb.get_sheet_by_name(sheets[1])
cell = data_sheet.cell(row = 22, column = 2)
resistance_cell = data_sheet.cell(row = 21, column = 4)
        
k = 1
j = 1
total_resistance = 0

while cell.value != None:
    i = 1
    total_resistance = 0 
    cell_past = cell
    while cell_past.value == cell.value:
        print("resistance cell ", resistance_cell)
        total_resistance += resistance_cell.value
        i += 1
        k += 1
        cell_past = cell
        resistance_cell = data_sheet.cell(row = (21 + k), column = 4)
        cell = data_sheet.cell(row = (21 + k), column = 2)

    
    average_resistance = total_resistance/i
    new_sheet.cell(row = j, column = 1). value = cell_past.value 
    new_sheet.cell(row = j, column = 2). value = average_resistance 
    j += 1
    print(average_resistance)

print("end of loop")
print(data_sheet.title)
wb.save(file_name)
